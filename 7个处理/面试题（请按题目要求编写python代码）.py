
import pandas
import os
######################## 笔试题

#1，循环读入data文件夹下所有文件（多个）
import pandas as pd

datas = []
files = os.listdir('data/')
print(files)
for file in os.listdir('data/'):
    datas.append(pd.read_excel('data/' + file))
data1 = datas[0]
data2 = datas[3]
#2，通过左连接的方式，连接data1 和data2,主键为合同号,新生产数据集为data_1_2下
data_1_2 = pd.merge(data1,data2,on='合同号',how='left')
data_1_2.to_excel('data/data_1_2.xlsx')



#3，通过循环的方式将所有表（data1...data11) 上下合并在一起，新生成的文件名称叫data_0
data_0 = pd.DataFrame()
for df in datas:
    data_0=data_0.append(df,ignore_index=False)




#4，对data_0中的变量 Down Payment 做分段处理，生成新的变量 Down Payment New（ 切分点为: <=0.1,[0.1,0.2),[0.2,0.3),>=0.3 )
bins = [-999999,0.1,0.2,0.3,999999]
score_cut = pd.cut(data_0['Down Payment'], bins)
df_groups = data_0.groupby(score_cut)
print(pd.value_counts(score_cut))

#5，对data_0中的合同号去重，取RPT_DATE最早的一条数据
data_0.drop_duplicates(['合同号'],keep='first')
data_0.to_excel('data/data_0.xlsx')

#6，使用数据data_0， 做如 “表格样例” 所示的表格（见《结果输出》）
l = data_0.values.tolist()
d = {}
for i in list(l):
    if str(i[0]).split()[0] not in d:
        d[str(i[0]).split()[0]] = [[],[],[],[]]
    if i[2] <= 0.1:
        d[str(i[0]).split()[0]][0].append(i[3])
    if 0.2 > i[2] >= 0.1:
        d[str(i[0]).split()[0]][1].append(i[3])
    if 0.3 > i[2] >= 0.2:
        d[str(i[0]).split()[0]][2].append(i[3])
    if i[2] > 0.3:
        d[str(i[0]).split()[0]][3].append(i[3])
res = []
for k,v in d.items():
    res.append([k,str(len(v[0])),str(sum(v[0])),str(len(v[1])),str(sum(v[1])),str(len(v[2])),str(sum(v[2])),str(len(v[3])),str(sum(v[3]))])
res.sort(key=lambda x:x[0])
df = pd.DataFrame(res,columns=['RPT_DATE','合同数','金额合计','合同数','金额合计','合同数','金额合计','合同数','金额合计'])



#7，将上题中生成的表格自动写入到Excel中的“数据输出区域”（见《结果输出》）
from openpyxl import load_workbook
book = load_workbook('结果输出.xlsx')
sheet = book.active
address = "C22"
start_row,start_col = sheet[address].row-1,sheet[address].column-1
for j, v in enumerate (df.columns,1):
    sheet.cell(start_row+1,start_col+j).value = v
for i, row in enumerate(df.values,2):
    for j, v in enumerate (row,1):
        sheet.cell(start_row+i,start_col+j).value = v
book.save("结果输出.xlsx")




