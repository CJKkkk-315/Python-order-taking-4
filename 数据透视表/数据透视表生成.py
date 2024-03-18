import pandas as pd
df1 = pd.read_excel('1.xlsx')
df2 = pd.read_excel('2.xlsx')
df1 = df1[df1['施工状态'] == '已完成']
df3 = pd.merge(df1, df2, on=['任务名称', '项目编码'])
df3.rename(columns={'项目名称_x':'项目名称'},inplace=True)
report = [0,0,0]

def count_rows(group):
    total_rows = len(group)
    completed_rows = len(group[group['施工状态'] == '已完成'])
    start_report_rows = len(group[group['开工报告提交日期'].notnull()])
    completion_report_rows = len(group[group['完工报告提交日期'].notnull()])
    report[0] += completed_rows
    report[1] += start_report_rows
    report[2] += completion_report_rows
    return pd.Series({'计数项:施工状态': completed_rows, '计数项:开工报告提交日期': start_report_rows,
                      '计数项:完工报告提交日期': completion_report_rows,'开工报告提交率': round(start_report_rows/completed_rows*100,2)
                         ,'完工报告提交率': round(completion_report_rows/completed_rows*100,2)})



# 按照项目名称和施工单位进行分组并计算统计量
df3_grouped_by_project_and_unit = df3.groupby(['项目名称', '施工单位']).apply(count_rows)


report = [0,0,0]

# 计算每个项目名称的总计，并设定与 df3_grouped_by_project_and_unit 相同的多级索引
df3_grouped_by_project = df3.groupby('项目名称').apply(count_rows).reset_index()
df3_grouped_by_project['施工单位'] = '总计'
df3_grouped_by_project.set_index(['项目名称', '施工单位'], inplace=True)
report.append(report[1]/report[0])
report.append(report[2]/report[0])
r = report[:]
print(r)
# 将每个项目名称的总计追加到各自的分组顶部
df3_final = pd.concat([df3_grouped_by_project, df3_grouped_by_project_and_unit]).reset_index()

# 创建一个新的辅助列，该列的值为施工单位是否为'总计'
df3_final['是否总计'] = df3_final['施工单位'] == '总计'

# 先按照是否为总计行排序（把总计行放在前面），然后再按照项目名称排序
df3_final.sort_values(by=['项目名称', '是否总计'], ascending=[True, False], inplace=True)

# 最后去掉辅助列，并重新设置索引
df3_final.drop(columns=['是否总计'], inplace=True)
df3_final.set_index(['项目名称', '施工单位'], inplace=True)
names = []
for i,j in df3_final.iterrows():
    if j['开工报告提交率'] != 100 or j['完工报告提交率'] != 100:
        names.append(list(j.name)[0])
report_str = f'截止目前，楚雄移动工程建设中心智慧工建已完工总数:{r[0]};PMS系统中开工报告提交数为{r[1]}，开工率{round(r[3]*100,2)}%;完工报告提交数为{r[2]}，完工率{round(r[4]*100,2)}%。\n开完工报告未达100%的项目是: {"，".join(names)}'
print(report_str)

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# 创建一个新的DataFrame，只包含一个单元格，其值为 "楚雄移动PMS工程项目开完工预警通报表"。
df_report = pd.DataFrame({report_str:[""]})
df_title = pd.DataFrame({"楚雄移动PMS工程项目开完工预警通报表": [""]})
df_subtitle = pd.DataFrame({"项目名称": [""],"施工单位": [""],"计数项:施工状态": [""],"计数项:开工报告提交日期": [""],"计数项:完工报告提交日期": [""],
                            "开工报告提交率": [""],"完工报告提交率": [""],})

# 使用 ExcelWriter 将 DataFrame 写入 Excel
with pd.ExcelWriter('output.xlsx', engine='openpyxl') as writer:
    df_report.to_excel(writer, startrow=0, index=False, header=True)
    df_title.to_excel(writer, startrow=1, index=False, header=True)
    df_subtitle.to_excel(writer, startrow=2, index=False, header=True)
    df3_final.to_excel(writer, startrow=3, index=True, header=False)

# 加载工作簿
book = load_workbook('output.xlsx')

# 获取工作表
sheet = book.active
sheet.title = 'PMS数据通报表'

# 合并A1到G1单元格
sheet.merge_cells('A1:G1')
sheet.merge_cells('A2:G2')

# 获取标题单元格
title_cell = sheet['A1']
font = Font(bold=False, name='微软雅黑')
for row in sheet.iter_rows(min_row=4):
    for cell in row:
        cell.font = font
title_cell.fill = PatternFill(start_color="C1FFC1",
                   end_color="C1FFC1",
                   fill_type="solid")
title_cell.font = Font(bold=True, name='微软雅黑')
from openpyxl.styles import Alignment
title_cell.alignment = Alignment(horizontal='left',wrap_text=True)
# 设置背景色为浅蓝色
fill = PatternFill(start_color="ADD8E6",
                   end_color="ADD8E6",
                   fill_type="solid")

# 设置字体为粗体
font = Font(bold=True, name='微软雅黑')

# 应用样式

sheet['A2'].fill = fill
sheet['A2'].font = font
sheet['A3'].fill = fill
sheet['A3'].font = font

sheet['B3'].fill = fill
sheet['B3'].font = font

sheet['C3'].fill = fill
sheet['C3'].font = font

sheet['D3'].fill = fill
sheet['D3'].font = font

sheet['E3'].fill = fill
sheet['E3'].font = font

sheet['F3'].fill = fill
sheet['F3'].font = font

sheet['G3'].fill = fill
sheet['G3'].font = font
for i in range(5, sheet.max_row):
    sheet['A' + str(i)].alignment = Alignment(vertical='center')
sheet.row_dimensions[1].height = 60
# 调整列宽
sheet.column_dimensions['A'].width = 70
sheet.column_dimensions['B'].width = 35
sheet.column_dimensions['C'].width = 15
sheet.column_dimensions['D'].width = 15
sheet.column_dimensions['E'].width = 15
sheet.column_dimensions['F'].width = 15
sheet.column_dimensions['G'].width = 15
from openpyxl.styles import Border, Side



# 保存修改后的工作簿
book.save('output.xlsx')

# 加载Excel文件
workbook = load_workbook('output.xlsx')
sheet = workbook.active

# 遍历每一行（从第2行开始，跳过表头）
for row in range(4,sheet.max_row+1):
    # 获取F列和G列的值
    f_value, g_value = sheet['F' + str(row)].value, sheet['G' + str(row)].value

    sheet['F' + str(row)].value = str(f_value) + '%'
    sheet['G' + str(row)].value = str(g_value) + '%'

    # 检查F列和G列的值是否小于100
    if f_value < 100:


        # 设置字体颜色为红色
        font_red = Font(color="FF0000", name='微软雅黑')
        sheet['F' + str(row)].font = font_red
        sheet['A' + str(row)].font = font_red
    if g_value < 100:


        # 设置字体颜色为红色
        font_red = Font(color="FF0000", name='微软雅黑')
        sheet['G' + str(row)].font = font_red
        sheet['A' + str(row)].font = font_red

# 获取C, D, E列的最后一行
last_row = sheet.max_row + 1

# 计算每一列的和并写入
report_all = []
for column in ['C', 'D', 'E']:
    # 计算列的和
    column_sum = r[ord(column) - ord('C')]

    # 获取最后一行的单元格
    cell = sheet[column + str(last_row)]

    # 写入计算的和
    cell.value = column_sum
    report_all.append(column_sum)
    # 设置样式为居中加粗
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True)

cell = sheet['F' + str(last_row)]
cell.value = str(round(r[3]*100,2)) + '%'
cell.font = Font(bold=True)

cell = sheet['G' + str(last_row)]
cell.value = str(round(r[4]*100,2)) + '%'
cell.font = Font(bold=True)


# 创建居中对齐样式
center_alignment = Alignment(horizontal='center')

# 应用对齐样式到C, D, E, F, G列的所有单元格
for column in ['C', 'D', 'E', 'F', 'G']:
    for cell in sheet[column]:
        cell.alignment = center_alignment
# 创建边框样式，所有边都为实线
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# 遍历所有单元格并应用边框样式
for row in sheet.iter_rows():
    for cell in row:
        cell.border = thin_border
# 保存修改后的工作簿
book.save('output.xlsx')

# 保存修改后的工作簿
book.save('output.xlsx')

new_sheet = workbook.create_sheet("智慧工建全量")
header = list(df3.columns)
new_sheet.append(header)
for _, row in df3.iterrows():
    new_sheet.append(list(row))



column_name = '开工报告完成时间'
filtered_df = df3[pd.isna(df3[column_name])]
new_sheet = workbook.create_sheet("开工未传")
header = list(filtered_df.columns)
new_sheet.append(header)
for _, row in filtered_df.iterrows():
    new_sheet.append(list(row))

column_name = '完工报告完成时间'
filtered_df = df3[pd.isna(df3[column_name])]
new_sheet = workbook.create_sheet("完工未传")
header = list(filtered_df.columns)
new_sheet.append(header)
for _, row in filtered_df.iterrows():
    new_sheet.append(list(row))

# 保存修改后的Excel文件
workbook.save('output.xlsx')