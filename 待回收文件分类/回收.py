import pandas as pd
from datetime import datetime
df = pd.read_excel('输入表格.xlsx')
xx = {}
zl = {}
df['未回收时长（天）'] = df['发送日期'].apply(lambda x:(datetime.today() - x).days)
for i,v in df.iterrows():
    if v['销售姓名'] not in xx:
        xx[v['销售姓名']] = []
        zl[v['销售姓名']] = {'交易确认书':0,'代签列表':0,'SAC协议、履约担保':0,'SAC协议':0,'代签列表、履约担保':0,'其他协议':0}
    xx[v['销售姓名']].append([v['编号'],str(v['发送日期']).split()[0].replace('-','/'),v['文件种类'],v['对手方名称/文件名称'],v['未回收时长（天）'],v['销售姓名']])
    zl[v['销售姓名']][v['文件种类']] += 1

from openpyxl import Workbook
from openpyxl.styles import Font, Color
for key in xx:
    s = f'截止{datetime.today().strftime("%Y年%m月%d日")}，您尚未回收的{"，".join([kk + ":" + str(zl[key][kk]) for kk in zl[key]])}。烦请尽快回收标红文件。'

    # 创建一个工作簿
    wb = Workbook()
    ws = wb.active
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['A'].width = 15
    # 定义红色字体
    red_font = Font(color=Color(rgb="00FF0000"))
    row_num = 1
    ws.append('编号	发送日期	文件种类	对手方名称/文件名称	未回收时长（天）	销售姓名'.split() + [s])
    bold_font = Font(bold=True)
    ws.cell(row=1,column=7).font = Font(bold=True)
    # 遍历列表，写入数据到工作簿中
    for row in xx[key]:
        ws.append(row)
        if row[4] > 30:
            ws.cell(row=row_num+1, column=5).font = red_font
        row_num += 1
    # 保存工作簿
    wb.save(f'{datetime.today().strftime("%Y%m%d") + key}.xlsx')
