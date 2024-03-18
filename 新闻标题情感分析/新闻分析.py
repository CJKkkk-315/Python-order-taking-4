import openpyxl
from snownlp import SnowNLP
import csv

# 定义情感分析函数
def sentiment_analysis(text):
    # 使用SnowNLP进行情感分析
    s = SnowNLP(text)
    # 得到情感分数
    sentiment = s.sentiments

    # 根据情感分数判断情感倾向：积极、消极或中性
    if sentiment >= 0.6:
        return "积极", sentiment
    elif sentiment < 0.4:
        return "消极", sentiment
    else:
        return "中性", sentiment


# 定义分析Excel文件的函数
def analyze_excel_file(file_path, text_column):
    # 加载Excel文件
    workbook = openpyxl.load_workbook(file_path)
    # 选择文件中的活动工作表
    sheet = workbook.active
    res_list = []
    # 遍历工作表的每一行，从第2行开始
    for row in range(2, sheet.max_row + 1):
        # 读取指定列的文本
        text = sheet.cell(row=row, column=text_column).value
        # 对文本进行情感分析
        res, sentiment = sentiment_analysis(text)
        # 打印情感分析结果
        print(f"第{row - 1}行：{res}, {sentiment}")
        res_list.append([res,sentiment])
    return res_list
# Excel文件路径
file_path = "评论exc(2).xlsx"
# 存储文本内容的列索引，例如A列为1，B列为2，以此类推
text_column = 1
# 调用analyze_excel_file函数进行情感分析
res_list = analyze_excel_file(file_path, text_column)
f = open('res.csv','w',newline='')
f_csv = csv.writer(f)
f_csv.writerows(res_list)