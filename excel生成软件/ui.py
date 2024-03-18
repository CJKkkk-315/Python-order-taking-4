import tkinter as tk
from tkinter import ttk
from tkinter import Entry, Label, Button
import tkinter.messagebox
def to_excel():
    import csv
    import openpyxl
    from openpyxl.styles import Alignment, Border, Side, PatternFill
    data = []
    with open('res.csv') as f:
        f_csv = csv.reader(f)
        for i in f_csv:
            data.append(i)
    print(data)
    # 创建一个工作簿
    wb = openpyxl.Workbook()
    ws = wb.active

    # 设置单元格边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 设置居中对齐
    center_aligned_text = Alignment(horizontal="center", vertical="center")
    fill_color = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')

    data.insert(0, [''])
    # 写入数据
    for i, row in enumerate(data):
        if len(row) == 1 and row[0] != '':
            # 只有一个单元格时，合并 A 到 Y
            merged_cell_range = f'A{i + 1}:Y{i + 1}'
            ws.merge_cells(merged_cell_range)
            cell = ws.cell(row=i + 1, column=1, value=row[0].strip())
            cell.alignment = center_aligned_text
            cell.fill = fill_color
            for col in range(1, 26):  # 为合并的单元格设置边框
                ws.cell(row=i + 1, column=col).border = thin_border
        else:
            for j, value in enumerate(row):
                # 去除前后空格
                cell_value = value.strip()
                try:
                    cell_value = str(round(float(cell_value), 3))
                except:
                    pass
                # 写入单元格
                cell = ws.cell(row=i + 1, column=j + 1, value=cell_value)
                # 设置边框和对齐
                if cell_value != '':
                    cell.border = thin_border
                    cell.alignment = center_aligned_text

    # 保存工作簿
    wb.save("res.xls")
def backend():

        all_cl = []
        all_excel = []
        city = a.get()

        if iii.get() == '':
            iii.insert(0, '0')
        if iii.get() == '0':
            disflag = False
        else:
            disflag = True
        # 东外墙
        city_map = """北京	0.0 
    天津	-0.1 
    沈阳	1.9 
    哈尔滨	-3.4 
    上海	0.5 
    南京	2.1 
    武汉	1.7 
    广州	0.0 
    昆明	-6.7 
    西安	0.9 
    兰州	-4.0 
    乌鲁木齐	0.2 
    重庆	2.0 """.split('\n')
        city_map = [i.split() for i in city_map]
        city_map = {i[0]:i[1] for i in city_map}
        time_0_23 = "0:00	1:00	2:00	3:00	4:00	5:00	6:00	7:00	8:00	9:00	10:00	11:00	12:00	13:00	14:00	15:00	16:00	17:00	18:00	19:00	20:00	21:00	22:00	23:00".split()
        dqtw1 = "38.5 	38.4 	38.2 	38.0 	37.6 	37.3 	36.9 	36.4 	36.0 	35.5 	35.2 	35.0 	35.0 	35.2 	35.6 	36.1 	36.6 	37.1 	37.5 	37.9 	38.2 	38.4 	38.5 	38.6".split()
        td = [city_map[city] for _ in range(len(time_0_23))]  # ia
        qka = "1.03	1.03	1.03	1.03	1.03	1.03	1.03	1.03	1.03	1.03	1.03	1.03	1.03	1.03	1.03	1.03	1.03	1.03	1.03	1.03	1.03	1.03	1.03	1.03".split()
        qkb = "0.94	0.94	0.94	0.94	0.94	0.94	0.94	0.94	0.94	0.94	0.94	0.94	0.94	0.94	0.94	0.94	0.94	0.94	0.94	0.94	0.94	0.94	0.94	0.94".split()
        dtwl = [(float(i1) + float(i2)) * float(i3) * float(i4) for i1, i2, i3, i4 in zip(dqtw1, td, qka, qkb)]
        tmp_in = [float(e.get()) for _ in range(len(time_0_23))]  # ie
        crk = [float(f.get()) for _ in range(len(time_0_23))]  # if
        dqarea = [float(iii.get()) for _ in range(len(time_0_23))]  # ii
        dcl = [float(i3) * float(i4) * (float(i1) - float(i2)) for i1, i2, i3, i4 in zip(dtwl, tmp_in, crk, dqarea)]
        dtwl = [round(i, 3) for i in dtwl]
        dcl = [round(i, 3) for i in dcl]
        wqtitle = """时间
            twl
            td
            kα
            kρ
            t'wl
            tNx
            K
            F
            CL""".split('\n')
        if disflag:
            all_cl.append(dcl)
            all_excel.append(['东外墙冷负荷'])
            all_excel.extend([time_0_23[:],dqtw1[:],td[:],qka[:],qkb[:],dtwl[:],tmp_in[:],crk[:],dqarea[:],dcl[:],[''],['']])
            for drr in range(len(wqtitle)):
                all_excel[-drr-3].insert(0,wqtitle[-drr-1])
        else:
            all_cl.append(['' for _ in range(len(time_0_23))])


        if jjj.get() == '':
            jjj.insert(0, '0')
        if jjj.get() == '0':
            disflag = False
        else:
            disflag = True

        # 南外墙
        city_map = """北京	0.0 
        天津	-0.4 
        沈阳	-1.4 
        哈尔滨	-2.2 
        上海	-0.8 
        南京	1.0 
        武汉	0.4 
        广州	-1.9 
        昆明	-8.5 
        西安	0.5 
        兰州	-4.8 
        乌鲁木齐	0.7 
        重庆	0.4 """.split('\n')
        city_map = [i.split() for i in city_map]
        city_map = {i[0]: i[1] for i in city_map}
        nqtw1 = "36.1 	36.2 	36.2 	36.1 	35.9 	35.6 	35.3 	35.0 	34.6 	34.2 	33.9 	33.5 	33.2 	32.9 	32.8 	32.9 	33.1 	33.4 	33.9 	34.4 	34.9 	35.3 	35.7 	36.0 ".split()
        td = [city_map[city] for _ in range(len(time_0_23))]
        ntwl = [(float(i1) + float(i2)) * float(i3) * float(i4) for i1, i2, i3, i4 in zip(nqtw1, td, qka, qkb)]
        nqarea = [float(jjj.get()) for _ in range(len(time_0_23))]  # ij
        ncl = [float(i3) * float(i4) * (float(i1) - float(i2)) for i1, i2, i3, i4 in zip(ntwl, tmp_in, crk, nqarea)]
        ntwl = [round(i, 3) for i in ntwl]
        ncl = [round(i, 3) for i in ncl]
        if disflag:
            all_cl.append(ncl)
            all_excel.append(['南外墙冷负荷'])
            all_excel.extend([time_0_23[:], nqtw1[:], td[:], qka[:], qkb[:], ntwl[:], tmp_in[:], crk[:], nqarea[:], ncl[:],[''],['']])
            for drr in range(len(wqtitle)):
                all_excel[-drr-3].insert(0,wqtitle[-drr-1])
        else:
            all_cl.append(['' for _ in range(len(time_0_23))])

        if k.get() == '':
            k.insert(0, '0')
        if k.get() == '0':
            disflag = False
        else:
            disflag = True
        # 西外墙
        city_map = """北京	0.0 
    天津	-0.1 
    沈阳	-1.9 
    哈尔滨	-3.4 
    上海	0.5 
    南京	2.1 
    武汉	1.7 
    广州	0.0 
    昆明	-6.7 
    西安	0.9 
    兰州	-4.0 
    乌鲁木齐	0.2 
    重庆	2.0""".split('\n')
        city_map = [i.split() for i in city_map]
        city_map = {i[0]: i[1] for i in city_map}
        xqtw1 = "38.5 	38.9 	39.1 	39.2 	39.1 	38.9 	38.6 	38.2 	37.8 	37.3 	36.8 	36.3 	35.9 	35.5 	35.2 	34.9 	34.8 	34.8 	34.9 	35.3 	35.8 	36.5 	37.3 	38.0 ".split()
        td = [city_map[city] for _ in range(len(time_0_23))]
        xtwl = [(float(i1) + float(i2)) * float(i3) * float(i4) for i1, i2, i3, i4 in zip(xqtw1, td, qka, qkb)]
        xqarea = [float(k.get()) for _ in range(len(time_0_23))]  # ik
        xcl = [float(i3) * float(i4) * (float(i1) - float(i2)) for i1, i2, i3, i4 in zip(xtwl, tmp_in, crk, xqarea)]
        xtwl = [round(i, 3) for i in xtwl]
        xcl = [round(i, 3) for i in xcl]
        if disflag:
            all_cl.append(xcl)
            all_excel.append(['西外墙冷负荷'])
            all_excel.extend([time_0_23[:], xqtw1[:], td[:], qka[:], qkb[:], xtwl[:], tmp_in[:], crk[:], xqarea[:], xcl[:],[''],['']])
            for drr in range(len(wqtitle)):
                all_excel[-drr-3].insert(0,wqtitle[-drr-1])
        else:
            all_cl.append(['' for _ in range(len(time_0_23))])

        if l.get() == '':
            l.insert(0, '0')
        if l.get() == '0':
            disflag = False
        else:
            disflag = True
        # 北外墙
        city_map = """北京	0.0 
    天津	-0.2 
    沈阳	-1.6 
    哈尔滨	-3.4 
    上海	1.2 
    南京	2.7 
    武汉	2.2 
    广州	1.7 
    昆明	-5.2 
    西安	1.8 
    兰州	-3.9 
    乌鲁木齐	-0.4 
    重庆	2.8 """.split('\n')
        city_map = [i.split() for i in city_map]
        city_map = {i[0]: i[1] for i in city_map}
        bqtw1 = "33.1 	33.2 	33.2 	33.2 	33.1 	33.0 	32.8 	32.6 	32.3 	32.1 	31.8 	31.6 	31.4 	31.3 	31.2 	31.2 	31.3 	31.4 	31.6 	31.8 	32.1 	32.4 	32.6 	32.9 ".split()
        td = [city_map[city] for _ in range(len(time_0_23))]
        btwl = [(float(i1) + float(i2)) * float(i3) * float(i4) for i1, i2, i3, i4 in zip(bqtw1, td, qka, qkb)]
        bqarea = [float(l.get()) for _ in range(len(time_0_23))]  # il
        bcl = [float(i3) * float(i4) * (float(i1) - float(i2)) for i1, i2, i3, i4 in zip(btwl, tmp_in, crk, bqarea)]
        btwl = [round(i, 3) for i in btwl]
        bcl = [round(i, 3) for i in bcl]
        if disflag:
            all_cl.append(bcl)
            all_excel.append(['北外墙冷负荷'])
            all_excel.extend([time_0_23[:], bqtw1[:], td[:], qka[:], qkb[:], btwl[:], tmp_in[:], crk[:], bqarea[:], bcl[:], [''], ['']])
            for drr in range(len(wqtitle)):
                all_excel[-drr-3].insert(0,wqtitle[-drr-1])
        else:
            all_cl.append(['' for _ in range(len(time_0_23))])


        if mmm.get() == '':
            mmm.insert(0, '0')
        if mmm.get() == '0':
            disflag = False
        else:
            disflag = True
        # 东内墙
        nqcrk = [float(g.get()) for _ in range(len(time_0_23))]  # ig
        dnqarea = [float(mmm.get()) for _ in range(len(time_0_23))]  # im
        twp = [float(c.get()) for _ in range(len(time_0_23))]  # ic
        c3 = [3 for _ in range(len(time_0_23))]
        tls = [i1 + i2 for i1, i2 in zip(twp, c3)]
        tmp_in = tmp_in
        dnqcl = [float(i1) * float(i2) * (float(i3) - float(i4)) for i1, i2, i3, i4 in zip(nqcrk, dnqarea, tls, tmp_in)]
        tls = [round(i, 3) for i in tls]
        dnqcl = [round(i, 3) for i in dnqcl]
        nqtitle = """时间
        K
        F
        twp
        △tls 
        tls
        tNx
        CL""".split('\n')
        if disflag:
            all_cl.append(dnqcl)
            all_excel.append(['东内墙冷负荷'])
            all_excel.extend([time_0_23[:], nqcrk[:], dnqarea[:], twp[:], c3[:], tls[:], tmp_in[:], dnqcl[:], [''], ['']])

            for drr in range(len(nqtitle)):
                all_excel[-drr-3].insert(0,nqtitle[-drr-1])
        else:
            all_cl.append(['' for _ in range(len(time_0_23))])

        if nnn.get() == '':
            nnn.insert(0, '0')
        if nnn.get() == '0':
            disflag = False
        else:
            disflag = True
        # 南内墙
        nnqarea = [float(nnn.get()) for _ in range(len(time_0_23))]  # in
        nnqcl = [float(i1) * float(i2) * (float(i3) - float(i4)) for i1, i2, i3, i4 in zip(nqcrk, nnqarea, tls, tmp_in)]
        nnqcl = [round(i, 3) for i in nnqcl]
        if disflag:
            all_cl.append(nnqcl)
            all_excel.append(['南内墙冷负荷'])
            all_excel.extend([time_0_23[:], nqcrk[:], nnqarea[:], twp[:], c3[:], tls[:], tmp_in[:], nnqcl[:], [''], ['']])
            for drr in range(len(nqtitle)):
                all_excel[-drr-3].insert(0,nqtitle[-drr-1])
        else:
            all_cl.append(['' for _ in range(len(time_0_23))])

        if ooo.get() == '':
            ooo.insert(0, '0')
        if ooo.get() == '0':
            disflag = False
        else:
            disflag = True
        # 西内墙
        xnqarea = [float(ooo.get()) for _ in range(len(time_0_23))]  # io
        xnqcl = [float(i1) * float(i2) * (float(i3) - float(i4)) for i1, i2, i3, i4 in zip(nqcrk, xnqarea, tls, tmp_in)]
        xnqcl = [round(i, 3) for i in xnqcl]
        if disflag:
            all_cl.append(xnqcl)
            all_excel.append(['西内墙冷负荷'])
            all_excel.extend([time_0_23[:], nqcrk[:], xnqarea[:], twp[:], c3[:], tls[:], tmp_in[:], xnqcl[:], [''], ['']])
            for drr in range(len(nqtitle)):
                all_excel[-drr-3].insert(0,nqtitle[-drr-1])
        else:
            all_cl.append(['' for _ in range(len(time_0_23))])

        if p.get() == '':
            p.insert(0, '0')
        if p.get() == '0':
            disflag = False
        else:
            disflag = True
        # 北内墙
        bnqarea = [float(p.get()) for _ in range(len(time_0_23))]  # ip
        bnqcl = [float(i1) * float(i2) * (float(i3) - float(i4)) for i1, i2, i3, i4 in zip(nqcrk, bnqarea, tls, tmp_in)]
        bnqcl = [round(i, 3) for i in bnqcl]
        if disflag:
            all_cl.append(bnqcl)
            all_excel.append(['北内墙冷负荷'])
            all_excel.extend([time_0_23[:], nqcrk[:], bnqarea[:], twp[:], c3[:], tls[:], tmp_in[:], bnqcl[:], [''], ['']])
            for drr in range(len(nqtitle)):
                all_excel[-drr-3].insert(0,nqtitle[-drr-1])
        else:
            all_cl.append(['' for _ in range(len(time_0_23))])


        if qqq.get() == '':
            qqq.insert(0, '0')
        if qqq.get() == '0':
            disflag = False
        else:
            disflag = True
        city = b.get()
        city_map = """北京	0.0 
    天津	0.0 
    石家庄	1.0 
    太原	-2.0 
    呼和浩特	-4.0 
    沈阳	-1.0 
    长春	-3.0 
    哈尔滨	-3.0 
    上海	1.0 
    南京	3.0 
    杭州	3.0 
    合肥	3.0 
    福州	2.0 
    南昌	3.0 
    济南	3.0 
    郑州	2.0 
    武汉	3.0 
    长沙	3.0 
    广州	1.0 
    南宁	1.0 
    成都	-1.0 
    贵阳	-3.0 
    昆明	-6.0 
    拉萨	-11.0 
    西安	2.0 
    兰州	-3.0 
    西宁	-8.0 
    银川	-3.0 
    乌鲁木齐	1.0 
    台北	1.0 
    二连	-2.0 
    汕头	1.0 
    海口	1.0 
    桂林	1.0 
    重庆	3.0 
    敦煌	-1.0 
    格尔木	-9.0 
    和田	-1.0 
    喀什	0.0 
    库车	0.0 """.split('\n')
        city_map = [i.split() for i in city_map if i]
        print(city_map)
        city_map = {i[0]: i[1] for i in city_map}
        # 东外窗瞬时
        tw1 = "27.20 	26.70 	26.2	25.80 	25.50 	25.3	25.4	26	26.9	27.9	29	29.9	30.80 	31.50 	31.90 	32.20 	32.20 	32.00 	31.60 	30.80 	29.90 	29.10 	28.40 	27.80 ".split()
        td = [float(city_map[city]) for _ in range(len(time_0_23))]  # ib
        tw1td = [float(i1) + i2 for i1, i2 in zip(tw1, td)]
        tmp_in = tmp_in
        cw_map = {
            '全部玻璃(单层)':1,
            '全部玻璃(双层)': 1,
            '木窗框80％玻璃(单层)': 0.9,
            '木窗框80％玻璃(双层)': 0.95,
            '木窗框60％玻璃(单层)': 0.8,
            '木窗框60％玻璃(双层)': 0.85,
            '金属窗框80％玻璃(单层)': 1,
            '金属窗框80％玻璃(双层)': 1.2,
        }
        cw = [cw_map[u.get()] for _ in range(len(time_0_23))]  # iu
        kcw = [float(h.get()) for _ in range(len(time_0_23))]  # ih
        dcarea = [float(qqq.get()) for _ in range(len(time_0_23))]  # iq
        dccl = [float(i1) * float(i2) * (float(i3) - float(i4)) for i1, i2, i3, i4 in zip(dcarea, kcw, tw1td, tmp_in)]
        dccl = [round(i, 3) for i in dccl]
        wctitle = """时间
        twl
        td
        twl+td
        tNx
        Cw
        Kcw
        Fcw
        CL""".split('\n')
        if disflag:
            all_cl.append(dccl)
            all_excel.append(['东外窗瞬时传热冷负荷'])
            all_excel.extend([time_0_23[:], tw1[:], td[:], tw1td[:], tmp_in[:], cw[:], kcw[:], dcarea[:], dccl[:], [''], ['']])

            for drr in range(len(wctitle)):
                all_excel[-drr-3].insert(0,wctitle[-drr-1])
        else:
            all_cl.append(['' for _ in range(len(time_0_23))])


        if r.get() == '':
            r.insert(0, '0')
        if r.get() == '0':
            disflag = False
        else:
            disflag = True
        # 南外窗瞬时
        ncarea = [float(r.get()) for _ in range(len(time_0_23))]  # ir
        nccl = [float(i1) * float(i2) * (float(i3) - float(i4)) for i1, i2, i3, i4 in zip(ncarea, kcw, tw1td, tmp_in)]
        nccl = [round(i, 3) for i in nccl]
        if disflag:
            all_cl.append(nccl)
            all_excel.append(['南外窗瞬时传热冷负荷'])
            all_excel.extend([time_0_23[:], tw1[:], td[:], tw1td[:], tmp_in[:],  cw[:], kcw[:], ncarea[:], nccl[:], [''], ['']])
            for drr in range(len(wctitle)):
                all_excel[-drr-3].insert(0,wctitle[-drr-1])
        else:
            all_cl.append(['' for _ in range(len(time_0_23))])


        if s.get() == '':
            s.insert(0, '0')
        if s.get() == '0':
            disflag = False
        else:
            disflag = True
        # 西外窗瞬时
        xcarea = [float(s.get()) for _ in range(len(time_0_23))]  # is
        xccl = [float(i1) * float(i2) * (float(i3) - float(i4)) for i1, i2, i3, i4 in zip(xcarea, kcw, tw1td, tmp_in)]
        xccl = [round(i, 3) for i in xccl]
        if disflag:
            all_cl.append(xccl)
            all_excel.append(['西外窗瞬时传热冷负荷'])
            all_excel.extend([time_0_23[:], tw1[:], td[:], tw1td[:], tmp_in[:],  cw[:], kcw[:], xcarea[:], xccl[:], [''], ['']])
            for drr in range(len(wctitle)):
                all_excel[-drr-3].insert(0,wctitle[-drr-1])
        else:
            all_cl.append(['' for _ in range(len(time_0_23))])


        if t.get() == '':
            t.insert(0, '0')
        if t.get() == '0':
            disflag = False
        else:
            disflag = True
        # 北外窗瞬时
        bcarea = [float(t.get()) for _ in range(len(time_0_23))]  # it
        bccl = [float(i1) * float(i2) * (float(i3) - float(i4)) for i1, i2, i3, i4 in zip(bcarea, kcw, tw1td, tmp_in)]
        bccl = [round(i, 3) for i in bccl]
        if disflag:
            all_cl.append(bccl)
            all_excel.append(['北外窗瞬时传热冷负荷'])
            all_excel.extend([time_0_23[:], tw1[:], td[:], tw1td[:], tmp_in[:],  cw[:], kcw[:], bcarea[:], bccl[:], [''], ['']])
            for drr in range(len(wctitle)):
                all_excel[-drr-3].insert(0,wctitle[-drr-1])
        else:
            all_cl.append(['' for _ in range(len(time_0_23))])

        if qqq.get() == '':
            qqq.insert(0, '0')
        if qqq.get() == '0':
            disflag = False
        else:
            disflag = True

        zytype = x.get()
        zy_map = """白布帘	0.50 
    浅蓝布帘	0.60 
    深黄、紫红、深绿布帘	0.65 
    活动百叶帘	0.60 
    无内遮阳	1.00""".split('\n')
        zy_map = [i.split() for i in zy_map]
        zy_map = {i[0]: i[1] for i in zy_map}
        city = b.get()
        city_idx_map = """海口
    福州	广州	南宁	贵阳	昆明	台北	汕头	桂林
    上海	南京	杭州	合肥	南昌	武汉	长沙	成都	重庆
    济南	郑州	西安	兰州	西宁	格尔木	和田
    北京	天津	石家庄	太原	呼和浩特	沈阳	银川	敦煌	喀什	库车
    长春	哈尔滨	乌鲁木齐	二连
    拉萨""".split('\n')
        djmax_i = -1
        city_idx_map = [i.split() for i in city_idx_map]
        for i in range(len(city_idx_map)):
            if city in city_idx_map[i]:
                djmax_i = i
                break
        djmax_v = """130 	311 	541 	465 	130 	465 	541 	311 	876 
    146 	332 	509 	421 	134 	421 	509 	332 	834 
    174 	374 	539 	415 	115 	415 	539 	374 	833 
    251 	436 	575 	430 	122 	430 	575 	436 	844 
    302 	477 	599 	442 	114 	442 	599 	477 	842 
    368 	508 	598 	432 	109 	432 	598 	508 	811 
    174 	462 	727 	592 	133 	593 	727 	462 	991 """.split('\n')
        djmax_v = [i.split() for i in djmax_v]
        ca_txt = """单层钢窗	单层木窗	双层钢窗	双层木窗
    0.85 	0.70 	0.75 	0.60""".split('\n')
        ca_txt = [i.split() for i in ca_txt]
        ca_map = {}
        for i in range(4):
            ca_map[ca_txt[0][i]] = ca_txt[1][i]
        cs_map = """标准玻璃	1.00 
    5mm厚普通玻璃	0.93 
    6mm厚普通玻璃	0.89 
    3mm厚吸热玻璃	0.96 
    5mm厚吸热玻璃	0.88 
    6mm厚吸热玻璃	0.83 
    双层3mm厚普通玻璃	0.86 
    双层5mm厚普通玻璃	0.78 
    双层6mm厚普通玻璃	0.74 """.split('\n')
        cs_map = [i.split() for i in cs_map]
        cs_map = {i[0]: i[1] for i in cs_map}
        bcity_all_l = """
                    贵阳						
    上海	南京	杭州	合肥	南昌	武汉	长沙	成都	重庆	
    济南	郑州	西安	兰州	西宁	格尔木	和田			
    北京	天津	石家庄	太原	呼和浩特	沈阳	银川	敦煌	喀什	库车
    长春	哈尔滨	乌鲁木齐	二连						
    拉萨									
    """.split('\n')
        bcity_all = []
        for bcity in bcity_all_l:
            for cc in bcity:
                bcity_all.append(cc)
        ncity_all = """福州	广州	南宁	昆明		台北	汕头	桂林""".split()
        if b.get() in bcity_all and zy_map[zytype] == 1:
            cqlv = "0.12 	0.11 	0.10 	0.09 	0.09 	0.08 	0.29 	0.41 	0.49 	0.60 	0.56 	0.37 	0.29 	0.29 	0.28 	0.26 	0.24 	0.22 	0.19 	0.17 	0.16 	0.15 	0.14 	0.13 ".split()
        elif b.get() in bcity_all and zy_map[zytype] != 1:
            cqlv = "0.06 	0.05 	0.05 	0.05 	0.04 	0.04 	0.47 	0.68 	0.82 	0.79 	0.59 	0.38 	0.24 	0.24 	0.23 	0.21 	0.18 	0.15 	0.11 	0.08 	0.07 	0.07 	0.06 	0.06 ".split()
        elif b.get() in ncity_all and zy_map[zytype] == 1:
            cqlv = "0.13 	0.11 	0.10 	0.09 	0.09 	0.08 	0.24 	0.39 	0.48 	0.61 	0.57 	0.38 	0.31 	0.30 	0.29 	0.28 	0.27 	0.23 	0.21 	0.18 	0.17 	0.15 	0.14 	0.13 ".split()
        else:
            cqlv = "0.06 	0.05 	0.05 	0.05 	0.04 	0.04 	0.36 	0.63 	0.81 	0.81 	0.63 	0.41 	0.27 	0.27 	0.25 	0.23 	0.20 	0.15 	0.10 	0.08 	0.07 	0.07 	0.07 	0.06 ".split()
        # 东窗日射
        cql = [float(cqlv[i]) for i in range(len(cqlv))]  # imore
        ca = [ca_map[v.get()] for _ in range(len(time_0_23))]  # iv
        djmax = [djmax_v[djmax_i][2] for _ in range(len(time_0_23))]  # ib
        cs = [cs_map[w.get()] for _ in range(len(time_0_23))]  # iw
        ci = [zy_map[zytype] for _ in range(len(time_0_23))]  # ix
        dcarea = dcarea
        dcrcl = [float(i1) * float(i2) * float(i3) * float(i4) * float(i5) * float(i6) for i1, i2, i3, i4, i5, i6 in
                 zip(cql, ca, djmax, cs, ci, dcarea)]
        wctitle = """时间
        CLQ
        Ca
        Dj,max
        Cs
        Ci
        Fcw
        CL""".split('\n')
        if disflag:
            all_cl.append(dcrcl)
            all_excel.append(['东窗日射得热形成的冷负荷'])
            all_excel.extend([time_0_23[:], cql[:], ca[:], djmax[:], cs[:], ci[:], dcarea[:], dcrcl[:], [''], ['']])

            for drr in range(len(wctitle)):
                all_excel[-drr-3].insert(0,wctitle[-drr-1])
        else:
            all_cl.append(['' for _ in range(len(time_0_23))])




        if b.get() in bcity_all and zy_map[zytype] == 1:
            cqlv = "0.16 	0.15 	0.14 	0.13 	0.12 	0.11 	0.13 	0.17 	0.21 	0.28 	0.39 	0.49 	0.54 	0.65 	0.60 	0.42 	0.36 	0.32 	0.27 	0.23 	0.21 	0.20 	0.18 	0.17 ".split()
        elif b.get() in bcity_all and zy_map[zytype] != 1:
            cqlv = "0.07 	0.07 	0.06 	0.06 	0.06 	0.05 	0.11 	0.18 	0.26 	0.40 	0.58 	0.72 	0.84 	0.80 	0.62 	0.45 	0.32 	0.24 	0.16 	0.10 	0.09 	0.09 	0.08 	0.08 ".split()
        elif b.get() in ncity_all and zy_map[zytype] == 1:
            cqlv = "0.21 	0.19 	0.18 	0.17 	0.16 	0.14 	0.17 	0.25 	0.33 	0.42 	0.48 	0.54 	0.59 	0.70 	0.70 	0.57 	0.52 	0.44 	0.35 	0.30 	0.28 	0.26 	0.24 	0.22 ".split()
        else:
            cqlv = "0.10 	0.09 	0.09 	0.08 	0.08 	0.07 	0.14 	0.31 	0.47 	0.60 	0.69 	0.77 	0.87 	0.84 	0.74 	0.66 	0.54 	0.38 	0.20 	0.13 	0.12 	0.12 	0.11 	0.10 ".split()

        if r.get() == '':
            s.insert(0, '0')
        if r.get() == '0':
            disflag = False
        else:
            disflag = True
        # 南窗日射
        cql = [float(cqlv[i]) for i in range(len(cqlv))]  # imore
        djmax = [djmax_v[djmax_i][2] for _ in range(len(time_0_23))]  # ib
        ncarea = ncarea
        ncrcl = [float(i1) * float(i2) * float(i3) * float(i4) * float(i5) * float(i6) for i1, i2, i3, i4, i5, i6 in
                 zip(cql, ca, djmax, cs, ci, ncarea)]
        if disflag:
            all_cl.append(ncrcl)
            all_excel.append(['南窗日射得热形成的冷负荷'])
            all_excel.extend([time_0_23[:], cql[:], ca[:], djmax[:], cs[:], ci[:], ncarea[:], ncrcl[:], [''], ['']])
            for drr in range(len(wctitle)):
                all_excel[-drr - 3].insert(0, wctitle[-drr - 1])
        else:
            all_cl.append(['' for _ in range(len(time_0_23))])



        if s.get() == '':
            s.insert(0, '0')
        if s.get() == '0':
            disflag = False
        else:
            disflag = True

        if b.get() in bcity_all and zy_map[zytype] == 1:
            cqlv = "0.17 	0.16 	0.15 	0.14 	0.13 	0.12 	0.12 	0.14 	0.15 	0.16 	0.17 	0.17 	0.18 	0.25 	0.37 	0.47 	0.52 	0.62 	0.55 	0.24 	0.23 	0.21 	0.20 	0.18 ".split()
        elif b.get() in bcity_all and zy_map[zytype] != 1:
            cqlv = "0.08 	0.07 	0.07 	0.06 	0.06 	0.06 	0.08 	0.11 	0.14 	0.17 	0.18 	0.19 	0.20 	0.34 	0.56 	0.72 	0.83 	0.77 	0.53 	0.11 	0.10 	0.09 	0.09 	0.08 ".split()
        elif b.get() in ncity_all and zy_map[zytype] == 1:
            cqlv = "0.17 	0.16 	0.15 	0.14 	0.13 	0.12 	0.12 	0.14 	0.16 	0.17 	0.18 	0.19 	0.20 	0.28 	0.40 	0.50 	0.54 	0.61 	0.50 	0.24 	0.23 	0.21 	0.20 	0.18 ".split()
        else:
            cqlv = "0.08 	0.07 	0.07 	0.06 	0.06 	0.06 	0.07 	0.12 	0.16 	0.19 	0.21 	0.22 	0.23 	0.37 	0.60 	0.75 	0.84 	0.73 	0.42 	0.10 	0.10 	0.09 	0.09 	0.08 ".split()
        # 西窗日射
        cql = [float(cqlv[i]) for i in range(len(cqlv))]  # imore
        djmax = [djmax_v[djmax_i][2] for _ in range(len(time_0_23))]  # ib
        xcarea = xcarea
        xcrcl = [float(i1) * float(i2) * float(i3) * float(i4) * float(i5) * float(i6) for i1, i2, i3, i4, i5, i6 in
                 zip(cql, ca, djmax, cs, ci, xcarea)]
        if disflag:
            all_cl.append(xcrcl)
            all_excel.append(['西窗日射得热形成的冷负荷'])
            all_excel.extend([time_0_23[:], cql[:], ca[:], djmax[:], cs[:], ci[:], xcarea[:], xcrcl[:], [''], ['']])
            for drr in range(len(wctitle)):
                all_excel[-drr - 3].insert(0, wctitle[-drr - 1])
        else:
            all_cl.append(['' for _ in range(len(time_0_23))])



        if b.get() in bcity_all and zy_map[zytype] == 1:
            cqlv = "0.26 	0.24 	0.23 	0.21 	0.19 	0.18 	0.44 	0.42 	0.43 	0.49 	0.56 	0.61 	0.64 	0.66 	0.66 	0.63 	0.59 	0.64 	0.64 	0.38 	0.35 	0.32 	0.30 	0.28 ".split()
        elif b.get() in bcity_all and zy_map[zytype] != 1:
            cqlv = "0.12 	0.11 	0.11 	0.10 	0.09 	0.09 	0.59 	0.54 	0.54 	0.65 	0.75 	0.81 	0.83 	0.83 	0.79 	0.71 	0.60 	0.61 	0.68 	0.17 	0.16 	0.15 	0.14 	0.13 ".split()
        elif b.get() in ncity_all and zy_map[zytype] == 1:
            cqlv = "0.28 	0.25 	0.24 	0.22 	0.21 	0.19 	0.38 	0.49 	0.52 	0.55 	0.59 	0.63 	0.66 	0.68 	0.68 	0.68 	0.69 	0.69 	0.60 	0.40 	0.37 	0.35 	0.32 	0.30 ".split()
        else:
            cqlv = "0.13 	0.12 	0.12 	0.11 	0.10 	0.10 	0.47 	0.67 	0.70 	0.72 	0.77 	0.82 	0.85 	0.84 	0.81 	0.78 	0.77 	0.75 	0.56 	0.18 	0.17 	0.16 	0.15 	0.14 ".split()


        if t.get() == '':
            t.insert(0, '0')
        if t.get() == '0':
            disflag = False
        else:
            disflag = True
        # 北窗日射
        cql = [float(cqlv[i]) for i in range(len(cqlv))]  # imore
        djmax = [djmax_v[djmax_i][2] for _ in range(len(time_0_23))]  # ib
        bcarea = bcarea
        bcrcl = [float(i1) * float(i2) * float(i3) * float(i4) * float(i5) * float(i6) for i1, i2, i3, i4, i5, i6 in
                 zip(cql, ca, djmax, cs, ci, bcarea)]
        bcrcl = [round(i, 3) for i in bcrcl]
        if disflag:
            all_cl.append(bcrcl)
            all_excel.append(['北窗日射得热形成的冷负荷'])
            all_excel.extend([time_0_23[:], cql[:], ca[:], djmax[:], cs[:], ci[:], bcarea[:], bcrcl[:], [''], ['']])
            for drr in range(len(wctitle)):
                all_excel[-drr-3].insert(0,wctitle[-drr-1])
        else:
            all_cl.append(['' for _ in range(len(time_0_23))])


        # 照明散热
        clq = "0.34	0.55	0.61	0.65	0.68	0.71	0.74	0.74	0.79	0.81	0.83	0.39	0.35	0.31	0.28	0.25	0.23	0.2	0.18	0.16	0.15	0.14	0.12	0.11".split()
        n1 = [1 for _ in range(len(time_0_23))]
        n2 = [0.6 for _ in range(len(time_0_23))]
        n = [float(y.get()) for _ in range(len(time_0_23))]  # iy
        m = [float(d.get())  for _ in range(len(time_0_23))]  # id
        zmcl = [float(i1) * float(i2) * float(i3) * float(i4) * float(i5) for i1, i2, i3, i4, i5 in zip(clq, n1, n2, n, m)]
        zmcl = [round(i, 3) for i in zmcl]
        all_cl.append(zmcl)
        all_excel.append(['照明散热形成的冷负荷'])
        all_excel.extend([time_0_23[:], clq[:], n1[:], n2[:], n[:], m[:], zmcl[:], [''], ['']])
        zmtitle = """时间
    C_LQ
    n1
    n2
    N
    M
    CL""".split('\n')
        for drr in range(len(zmtitle)):
            all_excel[-drr-3].insert(0,zmtitle[-drr-1])
        # 设备散热
        clq = "0.62 	0.69 	0.75 	0.79 	0.82 	0.84 	0.86 	0.88 	0.89 	0.91 	0.92 	0.93 	0.38 	0.31 	0.25 	0.21 	0.18 	0.16 	0.14 	0.12 	0.11 	0.09 	0.08 	0.07 ".split()
        n1 = [1 for _ in range(len(time_0_23))]
        n2 = [0.8 for _ in range(len(time_0_23))]
        n3 = [0.5 for _ in range(len(time_0_23))]
        sbcl = [float(i1) * float(i2) * float(i3) * float(i4) * float(i5) * float(i6) for i1, i2, i3, i4, i5, i6 in
                zip(clq, n1, n2, n3, n, m)]
        sbcl = [round(i, 3) for i in sbcl]
        all_cl.append(sbcl)
        all_excel.append(['设备散热形成的冷负荷'])
        all_excel.extend([time_0_23[:], clq[:], n1[:], n2[:], n3[:], n[:], m[:], sbcl[:], [''], ['']])
        sbtitle = """时间
    C_LQ
    n1
    n2
    n3
    N
    M
    CL""".split('\n')
        for drr in range(len(sbtitle)):
            all_excel[-drr-3].insert(0,sbtitle[-drr-1])
        tlhd_xr = {
            "静坐":"84	81	78	74	71	67	63	58	53	48	43".split(),
            "极轻劳动":"90	85	79	75	70	65	60.5	57	51	45	41".split(),
            "轻度劳动":"93	87	81	76	70	64	58	51	47	40	35".split(),
            "中等劳动":"117	112	104	97	88	83	74	67	61	52	45".split(),
            "重度劳动":"169	163	157	151	145	140	134	128	122	116	110".split()
        }

        tlhd_qr = {
            "静坐":"26	27	30	34	37	41	45	50	55	60	65".split(),
            "极轻劳动":"47	51	56	59	64	69	73.5	77	83	89	93".split(),
            "轻度劳动":"90	94	80	106	112	117	123	130	135	142	147".split(),
            "中等劳动":"118	123	131	138	147	152	161	168	174	183	190".split(),
            "重度劳动":"238	244	250	256	262	267	273	279	285	291	297".split()
        }
        qjxs = {
            "影剧院":0.89,
            "百货商店（售货）":0.89,
            "旅店":0.93,
            "体育馆":0.92,
            "图书阅览室":0.96,
            "工厂轻劳动":0.9,
            "银行":1.0,
            "工厂重劳动":1.0
        }
        # 人体散热
        clq = [0.55, 0.64, 0.7, 0.75, 0.79, 0.81, 0.84, 0.86, 0.88, 0.89, 0.91, 0.92, 0.45, 0.36, 0.3, 0.25, 0.21, 0.19,
               0.16, 0.14, 0.12, 0.11, 0.09, 0.08]
        q_s = [float(tlhd_xr[a1.get()][int(e.get())-20]) for _ in range(len(time_0_23))]  # ib1
        n = [int(c1.get()) for _ in range(len(time_0_23))]  # ic1
        o = [qjxs[d1.get()] for _ in range(len(time_0_23))]  # id1
        cls = [float(i1) * float(i2) * float(i3) * float(i4) for i1, i2, i3, i4 in zip(clq, q_s, n, o)]
        q_2 = [float(tlhd_qr[a1.get()][int(e.get())-20]) for _ in range(len(time_0_23))]  # ie
        q_t = [float(i1) * float(i2) * float(i3) for i1, i2, i3 in zip(n, o, q_2)]
        rtcl = [float(i1) + float(i2) for i1, i2 in zip(q_t, cls)]
        all_cl.append(rtcl)
        all_excel.append(['人体散热形成的冷负荷'])
        all_excel.extend([time_0_23[:], clq[:], q_s[:], n[:], o[:], cls[:], q_2[:], q_t[:], rtcl[:], [''], ['']])




        rttitle = """时间
    C_LQ
    q_s
    n
    φ
    CL_s
    q_2
    Q_τ
    合计""".split('\n')
        for drr in range(len(rttitle)):
            all_excel[-drr-3].insert(0,rttitle[-drr-1])
        # 新风冷负荷计算
        hn = float(g1.get())  # ig1
        hw = float(f1.get())  # if1
        m0 = float(e1.get())  # ie1
        q = m0 * 1.2 / 3600 * (hw - hn) * 1000
        all_cl.append([q for _ in range(len(time_0_23))])
        all_excel.append(['新风冷负荷计算'])
        all_excel.extend([[hn for _ in range(len(time_0_23))], [hw for _ in range(len(time_0_23))], [m0 for _ in range(len(time_0_23))], [q for _ in range(len(time_0_23))], [''], ['']])
        xfltitle = """hn
    hw
    M0
    Q""".split('\n')
        for drr in range(len(xfltitle)):
            all_excel[-drr-3].insert(0,xfltitle[-drr-1])


        res_cl = []
        for i in range(len(time_0_23)):
            now = 0
            for j in all_cl:
                if j[i] != '':
                    now += float(j[i])
            res_cl.append(now)
        all_excel.append(['各分项逐时冷负荷汇总表'])
        alll_cl = [time_0_23] + all_cl + [res_cl]
        print(alll_cl)
        hztitle = """时间
    东外墙冷负荷
    南外墙冷负荷
    西外墙冷负荷
    北外墙冷负荷
    东内墙冷负荷
    南内墙冷负荷
    西内墙冷负荷
    北内墙冷负荷
    东窗传热冷负荷
    南窗传热冷负荷
    西窗传热冷负荷
    北窗传热冷负荷
    东窗透热冷负荷
    南窗透热冷负荷
    西窗透热冷负荷
    北窗透热冷负荷
    照明散热冷负荷
    设备散热冷负荷
    人体散热冷负荷
    新风冷负荷
    合计""".split('\n')
        for drr in range(len(hztitle)):
            alll_cl[-drr-1].insert(0,hztitle[-drr-1])
        clean_all_cl = []
        for clcl in alll_cl:
            if clcl[1] != '':
                clean_all_cl.append(clcl)
        all_excel.extend(clean_all_cl)
        import csv
        with open('res.csv','w',newline='') as file:
            f_csv = csv.writer(file)
            f_csv.writerows(all_excel)

        to_excel()
        tk.messagebox.showinfo('结果','生成成功！')

root = tk.Tk()
# root.title("界面仿真软件")
root.geometry("1450x700")
for i in range(12):
    root.grid_columnconfigure(i, uniform="column")
from tkinter.font import Font
font = Font(size = 9)
root.option_add("*TCombobox*Listbox*Font", font)


Label(root, text=" 外墙修正\n城市td:").grid(row=0, column=0)
city_map = """北京	0.0 
天津	-0.1 
沈阳	1.9 
哈尔滨	-3.4 
上海	0.5 
南京	2.1 
武汉	1.7 
广州	0.0 
昆明	-6.7 
西安	0.9 
兰州	-4.0 
乌鲁木齐	0.2 
重庆	2.0 """.split('\n')
citys = [i.split()[0] for i in city_map]
a = ttk.Combobox(root,width=5)
a['values'] = citys
a.grid(row=0, column=1)

city_map = """北京	0.0 
天津	0.0 
石家庄	1.0 
太原	-2.0 
呼和浩特	-4.0 
沈阳	-1.0 
长春	-3.0 
哈尔滨	-3.0 
上海	1.0 
南京	3.0 
杭州	3.0 
合肥	3.0 
福州	2.0 
南昌	3.0 
济南	3.0 
郑州	2.0 
武汉	3.0 
长沙	3.0 
广州	1.0 
南宁	1.0 
成都	-1.0 
贵阳	-3.0 
昆明	-6.0 
拉萨	-11.0 
西安	2.0 
兰州	-3.0 
西宁	-8.0 
银川	-3.0 
乌鲁木齐	1.0 
台北	1.0 
二连	-2.0 
汕头	1.0 
海口	1.0 
桂林	1.0 
重庆	3.0 
敦煌	-1.0 
格尔木	-9.0 
和田	-1.0 
喀什	0.0 
库车	0.0 """.split('\n')
citys = [i.split()[0] for i in city_map if i]
Label(root, text="外窗修正城市:").grid(row=0, column=2)
b = ttk.Combobox(root,width=5)
b['values'] = citys
b.grid(row=0, column=3)

Label(root, text="夏季室外计算\n日平均温度:").grid(row=0, column=4)
c = Entry(root, width=5)
c.grid(row=0, column=5)

Label(root, text="房间面积m²:").grid(row=0, column=6)
d = Entry(root, width=5)
d.grid(row=0, column=7)


Label(root, text="室内温度°C:").grid(row=1, column=0)
e = Entry(root, width=5)
e.grid(row=1, column=1)

now_row = 2

Label(root, text="传热系数\n(W/(m2.K)):").grid(row=now_row, column=1)

Label(root, text="东向(m²):").grid(row=now_row, column=3)

Label(root, text="南向(m²)").grid(row=now_row, column=5)

Label(root, text="西向(m²)").grid(row=now_row, column=7)

Label(root, text="北向(m²)").grid(row=now_row, column=9)

Label(root, text="外墙").grid(row=now_row, column=0)



f = Entry(root, width=5)
f.grid(row=now_row, column=2)

iii = Entry(root, width=5)
iii.grid(row=now_row, column=4)

jjj = Entry(root, width=5)
jjj.grid(row=now_row, column=6)

k = Entry(root, width=5)
k.grid(row=now_row, column=8)

l = Entry(root, width=5)
l.grid(row=now_row, column=10)


now_row = 3

Label(root, text="传热系数\n(W/(m2.K)):").grid(row=now_row, column=1)

Label(root, text="东向(m²):").grid(row=now_row, column=3)

Label(root, text="南向(m²)").grid(row=now_row, column=5)

Label(root, text="西向(m²)").grid(row=now_row, column=7)

Label(root, text="北向(m²)").grid(row=now_row, column=9)


Label(root, text="内墙").grid(row=now_row, column=0)



g = Entry(root, width=5)
g.grid(row=now_row, column=2)

mmm = Entry(root, width=5)
mmm.grid(row=now_row, column=4)

nnn = Entry(root, width=5)
nnn.grid(row=now_row, column=6)

ooo = Entry(root, width=5)
ooo.grid(row=now_row, column=8)

p = Entry(root, width=5)
p.grid(row=now_row, column=10)


now_row = 4


Label(root, text="传热系数\n(W/(m2.K)):").grid(row=now_row, column=1)

Label(root, text="东向(m²):").grid(row=now_row, column=3)

Label(root, text="南向(m²)").grid(row=now_row, column=5)

Label(root, text="西向(m²)").grid(row=now_row, column=7)

Label(root, text="北向(m²)").grid(row=now_row, column=9)

Label(root, text="外墙").grid(row=now_row, column=0)

Label(root, text="内墙").grid(row=3, column=0)

Label(root, text="外窗").grid(row=4, column=0)


h = Entry(root, width=5)
h.grid(row=now_row, column=2)

qqq = Entry(root, width=5)
qqq.grid(row=now_row, column=4)

r = Entry(root, width=5)
r.grid(row=now_row, column=6)

s = Entry(root, width=5)
s.grid(row=now_row, column=8)

t = Entry(root, width=5)
t.grid(row=now_row, column=10)

now_row = 5
Label(root, text="外窗参数").grid(row=now_row, column=0)

cw_map = {
        '全部玻璃(单层)':1,
        '全部玻璃(双层)': 1,
        '木窗框80％玻璃(单层)': 0.9,
        '木窗框80％玻璃(双层)': 0.95,
        '木窗框60％玻璃(单层)': 0.8,
        '木窗框60％玻璃(双层)': 0.85,
        '金属窗框80％玻璃(单层)': 1,
        '金属窗框80％玻璃(双层)': 1.2,
    }
Label(root, text="窗框类型cw:").grid(row=now_row, column=1)
u = ttk.Combobox(root, width=16)
u['values'] = list(cw_map.keys())
u.grid(row=now_row, column=2)

ca_txt = """单层钢窗	单层木窗	双层钢窗	双层木窗
0.85 	0.70 	0.75 	0.60""".split('\n')
ca_txt = [i.split() for i in ca_txt]
ca_map = {}
for i in range(4):
    ca_map[ca_txt[0][i]] = ca_txt[1][i]
Label(root, text="窗有效面积\n系数ca:").grid(row=now_row, column=3)
v = ttk.Combobox(root, width=12)
v['values'] = list(ca_map.keys())
v.grid(row=now_row, column=4)

cs_map = """标准玻璃	1.00 
5mm厚普通玻璃	0.93 
6mm厚普通玻璃	0.89 
3mm厚吸热玻璃	0.96 
5mm厚吸热玻璃	0.88 
6mm厚吸热玻璃	0.83 
双层3mm厚普通玻璃	0.86 
双层5mm厚普通玻璃	0.78 
双层6mm厚普通玻璃	0.74 """.split('\n')
cs_map = [i.split() for i in cs_map]
cs_map = {i[0]: i[1] for i in cs_map}
Label(root, text="玻璃窗\n系数cs:").grid(row=now_row, column=5)
w = ttk.Combobox(root, width=12)
w['values'] = list(cs_map.keys())
w.grid(row=now_row, column=6)


zy_map = """白布帘	0.50 
浅蓝布帘	0.60 
深黄、紫红、深绿布帘	0.65 
活动百叶帘	0.60 
无内遮阳	1.00""".split('\n')
zy_map = [i.split() for i in zy_map]
zy_map = {i[0]: i[1] for i in zy_map}
Label(root, text="内遮阳\n系数ci:").grid(row=now_row, column=7)
x = ttk.Combobox(root, width=12)
x['values'] = list(zy_map.keys())
x.grid(row=now_row, column=8)

now_row = 6

Label(root, text="照明功率\n密度w/m²:").grid(row=now_row, column=1)
y = Entry(root, width=5)
y.grid(row=now_row, column=2)

Label(root, text="设备功率\n密度:").grid(row=now_row, column=3)
z = Entry(root, width=5)
z.grid(row=now_row, column=4)

now_row = 7

tlhd_qr = {
    "静坐": "26	27	30	34	37	41	45	50	55	60	65".split(),
    "极轻劳动": "47	51	56	59	64	69	73.5	77	83	89	93".split(),
    "轻度劳动": "90	94	80	106	112	117	123	130	135	142	147".split(),
    "中等劳动": "118	123	131	138	147	152	161	168	174	183	190".split(),
    "重度劳动": "238	244	250	256	262	267	273	279	285	291	297".split()
}
Label(root, text="体力活动\n性质:").grid(row=now_row, column=1)
a1 = ttk.Combobox(root, width=8)
a1['values'] = list(tlhd_qr)
a1.grid(row=now_row, column=2)

Label(root, text="场所:").grid(row=now_row, column=3)
b1 = ttk.Combobox(root, width=12)
changsuo = """
"影剧院    会堂
阅览室"



"旅馆       体育馆
手表装配    电子元件"



"百货商店    化学实验室
手表装配    电子元件"



"纺织车间    印刷车间
机加工车间    "



"炼钢车间  铸造车间
排练厅               室内运动场   "
""".replace('"','').split('\n')
changsuo = [i.split() for i in changsuo if i]

chansguos = []
for i in changsuo:
    for j in i:
        chansguos.append(j)

b1['values'] = chansguos
b1.grid(row=now_row, column=4)

Label(root, text="人数:").grid(row=now_row, column=5)
c1 = Entry(root, width=5)
c1.grid(row=now_row, column=6)


qjxs = {
        "影剧院":0.89,
        "百货商店（售货）":0.89,
        "旅店":0.93,
        "体育馆":0.92,
        "图书阅览室":0.96,
        "工厂轻劳动":0.9,
        "银行":1.0,
        "工厂重劳动":1.0
    }
Label(root, text="群集系数:").grid(row=now_row, column=7)
d1 = ttk.Combobox(root, width=12)
d1['values'] = list(qjxs.keys())
d1.grid(row=now_row, column=8)

Label(root, text="人在室内\n总小时数:").grid(row=now_row, column=9)
d2 = Entry(root, width=5)
d2.grid(row=now_row, column=10)

now_row = 8

Label(root, text="新风量m³/h:").grid(row=now_row, column=1)
e1 = Entry(root, width=5)
e1.grid(row=now_row, column=2)

Label(root, text="室外焓值\nhw KJ/kg:").grid(row=now_row, column=3)
f1 = Entry(root, width=5)
f1.grid(row=now_row, column=4)

Label(root, text="室内焓值\nhw KJ/kg:").grid(row=now_row, column=5)
g1 = Entry(root, width=5)
g1.grid(row=now_row, column=6)

Label(root, text="室内焓值\nhw KJ/kg:").grid(row=now_row, column=5)
Button(root, text='生成', command=backend).grid(row=now_row, column=7)
Label(root, text="""
注   1、墙体默认采用Ⅱ型外墙参数。
       2、外墙修正地点的选取是根据《空调工程》附录9选取
       3、外窗修正地点的选取是根据《空调工程》附录15选取
       4、如果设计城市在附录中不包含，就近选择城市即可
       5、由于外窗修正城市比外墙修正城市多，所以预留两个窗口，计算结果更准确
       6、如果该方向朝向没有墙体，框内添加0或者不添加均可
       7、内墙如果需要即可添加，不需要可添加0或者不添加
       8、外墙面积需要减掉外窗面积
""").place(x=500,y=400)
root.mainloop()
