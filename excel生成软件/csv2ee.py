import csv
import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill
data = []
with open('res.csv') as f:
    f_csv = csv.reader(f)
    for i in f_csv:
        data.append(i)
print(data)
# 创建一个工作簿
wb = openpyxl.Workbook()
ws = wb.active

# 设置单元格边框样式
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 设置居中对齐
center_aligned_text = Alignment(horizontal="center", vertical="center")
fill_color = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')

data.insert(0,[''])
# 写入数据
for i, row in enumerate(data):
    if len(row) == 1 and row[0] != '':
        # 只有一个单元格时，合并 A 到 Y
        merged_cell_range = f'A{i + 1}:Y{i + 1}'
        ws.merge_cells(merged_cell_range)
        cell = ws.cell(row=i + 1, column=1, value=row[0].strip())
        cell.alignment = center_aligned_text
        cell.fill = fill_color
        for col in range(1, 26):  # 为合并的单元格设置边框
            ws.cell(row=i + 1, column=col).border = thin_border
    else:
        for j, value in enumerate(row):
            # 去除前后空格
            cell_value = value.strip()
            try:
                cell_value = str(round(float(cell_value),3))
            except:
                pass
            # 写入单元格
            cell = ws.cell(row=i+1, column=j+1, value=cell_value)
            # 设置边框和对齐
            if cell_value != '':
                cell.border = thin_border
                cell.alignment = center_aligned_text

# 保存工作簿
wb.save("output.xls")