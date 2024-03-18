import numpy as np
import matplotlib.pyplot as plt
#本任务中并不要求必须使用中文文字
#如需使用中文，需要使用下面的语句将中文字体设置为黑体
#如果系统原因显示不出来中文，无需纠结，英文亦可
plt.rcParams["font.sans-serif"]=["SimHei"]

#-------下划线和省略号处需要替换成相应的代码

#*****任务3-Section 1-Begin
import openpyxl#导入包

#通过load_workbook打开excel
wb = openpyxl.load_workbook('yututwo.xlsx')

#wb是一个字典，获取一个Sheet对象
sheet = wb['Sheet1']

#初始化两个列表，用于存储表1中的数据
#为了便于标签显示，建议年列表设计为字符串列表
years = []
distance = []

for i in range(2,sheet.max_row+1):
    years.append(str(sheet['A'+str(i)].value))
    distance.append(float(sheet['B'+str(i)].value))

#*****任务3-Section1-End

#任务2：计算累计行驶距离，存储在另一个列表中
reduce_dis = []
for i in range(1, len(years)+1):
    reduce_dis.append(sum(distance[:i]))

#*****任务3-Section2-Begin
#将累计行驶距离写入C列
sheet['C1'] = '累计行驶距离（米）' #列标题写入C1单元格

#循环写入C列其他单元格
for i in range(2,sheet.max_row):
    sheet['C'+str(i)] = reduce_dis[i-2]

#保存excel文件
wb.save('yututwo.xlsx')

#*****任务3-Section2-End


#绘制饼图和折线图，figsize用于设置图片尺寸，建议保留
plt.figure(figsize=(12, 6))

#第一个子图绘制饼图
plt.subplot(121)
plt.pie(
    distance ,  #此处填写年行驶距离列表
    labels=years, #此处填写年列表，作为图的标签
    autopct='%.2f%%', #饼图上显示的比例值，保留两位小数的百分数
)
plt.title('月兔二号行驶距离比例') #此处填写饼图的Title

#第二个子图绘制折线图
plt.subplot(122)
plt.plot(years,reduce_dis,marker='o')
plt.xlabel('年')
plt.ylabel('累计行驶距离(米)')
plt.title('月兔二号累计行驶距离')
#保存图
plt.savefig('22379197-3.png')

#显示统计图
plt.show()
