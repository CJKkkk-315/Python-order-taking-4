import openpyxl


def append_data(file_path, data):
    # 加载工作簿
    workbook = openpyxl.load_workbook(file_path)

    # 选择工作表
    worksheet = workbook.active

    # 找到最后一行
    last_row = worksheet.max_row

    # 添加新数据
    for i, row_data in enumerate(data):
        for j, cell_data in enumerate(row_data):
            worksheet.cell(row=last_row + i + 1, column=j + 1, value=cell_data)

    # 保存工作簿
    workbook.save(file_path)


# 使用示例
append_data(
    "副本CRM_客户_导入模板 (1)(9)(3)(13).xlsx",  # Excel文件路径
    [[11, 12, 13], [21, 22, 23], [31, 32, 33]]  # 要插入的数据
)
