import requests
import json
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import os
xlsx_file = [i for i in os.listdir() if i[-4:] == 'xlsx'][0]
res = []
def get_unread():
    all_info = []
    url = "https://partner.zmlearn.com/api/zmbiz-sale-message/mc/page/saleMsg"
    payload = '{' + f'"state":"1","sortField":"createdAt","sortType":"DESC","titleType":null,"detailType":null,"keyword":"","pageSize":100,"pageNo":{1}' + '}'

    response = requests.request("POST", url, headers=headers1, data=payload)
    # print(response.text)
    page_size = json.loads(response.text)['data']['pages']
    for page in range(1, page_size+1):
        payload = '{' + f'"state":"1","sortField":"createdAt","sortType":"DESC","titleType":null,"detailType":null,"keyword":"","pageSize":100,"pageNo":{page}' + '}'
        response = requests.request("POST", url, headers=headers1, data=payload)
        rows = json.loads(response.text)['data']['list']
        for row in rows:
            all_info.append(row['content'])
    return all_info
def unread2phone(all_info):
    phones = set()
    for info in all_info:
        pattern = re.compile(r'【学生手机号：(\d+)】')
        match = pattern.search(info)
        if match:
            phone = match.group(1)
            phones.add(phone)
    return phones
def phone2detail(phones):
    phones = list(phones)
    url1 = "https://partner.zmlearn.com/api/student-tag/student/queryForPage"
    url2 = "https://partner.zmlearn.com/api/student/student/getStudentActivityLogList"
    for phone in phones:
        try:
            payload = '{' + f'"stuMobile":"","sortField":"responseTime","sortType":"desc","pageNo":1,"pageSize":10,"state":"","tmkRemoveList":false,"isImport":false,"isTodayTodo":false,"studentName":"","is7DaysUnContacted":false,"is14DaysUnTranslated":false,"is14DaysUnTranslatedFinished":false,"isPlanTodayFinished":false,"isTrialType":false,"todoNameList":[],"accountNumber":"{phone}","tagIds":[],"buId":102,"positionCode":"CR_LEADER"' + '}'
            response = requests.request("POST", url1, headers=headers2, data=payload)
            all_stu = json.loads(response.text)['data']['list']

            if len(all_stu) == 0:
                name = ''
                city = ''
                year = ''
                provice = ''
                mark = ''
                date = ''
                res.append(['',name,phone,'','','',year,'',provice,city,'','','',mark,'','',date])
                continue
            stuinfo = all_stu[0]
            name = stuinfo['studentName']
            city = stuinfo['city']
            grade = stuinfo['grade']
            provice = stuinfo['provice']
            stukey = stuinfo['stuKey']
            stuid = stuinfo['studentId']
            payload = '{' + f'"pageNo":{1},"pageSize":10,"sellerRole":null,"endDate":null,"startDate":null,"stuKey":"{stukey}"' + '}'
            response = requests.request("POST", url2, headers=headers2, data=payload)
            page_size = json.loads(response.text)['data']['pages']
            payload = '{' + f'"pageNo":{page_size},"pageSize":10,"sellerRole":null,"endDate":null,"startDate":null,"stuKey":"{stukey}"' + '}'
            response = requests.request("POST", url2, headers=headers2, data=payload)
            year = g2y.get(grade,'')
            if len(json.loads(response.text)['data']['list']) != 0:
                mark = json.loads(response.text)['data']['list'][-1]['mark']
                date = json.loads(response.text)['data']['list'][-1]['operateDate']
            else:
                mark = ''
                date = ''
            res.append(['', name, phone, '', '', '', year, '', provice, city, '', '', '', mark, '', '', date])
            print([name, phone, provice, city, grade, mark, date])
            # print(headers2)
            url = "https://partner.zmlearn.com/api/zmbiz-sale-comment/access/saveAccess"
            payload = '{' + f'"stuId":{stuid},"context":"跟进","warnFlage":true,"version":"3.0","groupId":"","tagIds":[10],"tagNames":["常规回访"]' + '}'
            response = requests.request("POST", url, headers=headers2, data=payload.encode('utf-8'))
            # print(response.text)
        except:
            continue
def append_data(file_path, data):
    # 加载工作簿
    workbook = openpyxl.load_workbook(file_path)

    # 选择工作表
    worksheet = workbook.active

    # 找到最后一行
    last_row = worksheet.max_row
    while last_row > 0:
        if any(cell.value for cell in worksheet[last_row]):
            break  # 找到第一个非空行，跳出循环
        last_row -= 1  # 如果这一行是空的，检查上一行
    # 添加新数据
    for i, row_data in enumerate(data):
        for j, cell_data in enumerate(row_data):
            worksheet.cell(row=last_row + i + 1, column=j + 1, value=cell_data)

    # 保存工作簿
    workbook.save(file_path)
g2y = """2024高四
2024高三
2025高二
2026高一
2027初三
2028初二
2029初一
2030小六
2031小五
2032小四
2033小三
2034小二
2035小一""".split('\n')
g2y = {i[-2:]:i[:4] for i in g2y}
user_info = [i.split('----') for i in open('账号信息.txt').read().split('\n') if i]
for user in user_info:
        username = user[0]
        password = user[1]
        driver = webdriver.Chrome(executable_path='chromedriver.exe')
        driver.get('https://partner.zmlearn.com/#/login')
        WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="app"]/div/form/div[1]/div/div[1]/input'))
        )
        driver.find_element(by=By.XPATH, value='//*[@id="app"]/div/form/div[1]/div/div[1]/input').send_keys(username)
        driver.find_element(by=By.XPATH, value='//*[@id="app"]/div/form/div[3]/div/div[1]/input').send_keys(password)
        driver.find_element(by=By.XPATH, value='//*[@id="app"]/div/form/div[5]/div/button[1]').click()
        WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="app"]/div/ul/div[2]/a'))
        )
        driver.find_element(by=By.XPATH, value='//*[@id="app"]/div/ul/div[2]/a').click()
        cookies = driver.get_cookies()
        acw_tc = cookies[1]['value']
        driver.quit()
        PARTNER_TOKEN = cookies[0]['value']
        headers1 = {
              'Cookie': f'acw_tc={acw_tc}; PARTNER_TOKEN={PARTNER_TOKEN};',
              'Content-Type': 'text/plain',
            'Connection': 'close'
            }

        headers2 = {
              'Cookie': f'acw_tc={acw_tc}; PARTNER_TOKEN={PARTNER_TOKEN};',
               'Content-Type': 'application/json;charset=UTF-8',
                'Connection': 'close'
            }
        all_info = get_unread()
        phones = unread2phone(all_info)
        phone2detail(phones)
        url3 = 'https://partner.zmlearn.com/api/zmbiz-sale-message/mc/readLimit'
        payload = '{' + '"typeList":[0,5]' + '}'
        response = requests.request("POST", url3, headers=headers2, data=payload)
        print(response.text)
        append_data(xlsx_file, res)
        res = []


